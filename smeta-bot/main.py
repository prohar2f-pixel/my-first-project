import os
import io
import json
import asyncio
import logging
import httpx
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from telegram import Update, BotCommand
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
OPENROUTER_API_KEY = os.environ["OPENROUTER_API_KEY"]
SERPER_API_KEY = os.environ["SERPER_API_KEY"]

logger.info(f"OpenRouter API Key loaded: {OPENROUTER_API_KEY[:20]}... (length: {len(OPENROUTER_API_KEY)})")

OPENROUTER_HEADERS = {
    "Authorization": f"Bearer {OPENROUTER_API_KEY}",
    "HTTP-Referer": "smeta-bot",
    "X-Title": "Smeta Bot",
    "Content-Type": "application/json"
}

logger.info(f"OpenRouter Headers configured")
logger.info(f"OpenRouter API Base URL: https://openrouter.ai/api/v1/chat/completions")

cancelled_chats: set[int] = set()

DEFAULT_REGION = "Москва"

RELIABILITY_STARS = {"высокая": "⭐⭐⭐⭐⭐", "средняя": "⭐⭐⭐", "низкая": "⭐"}


def search_suppliers(material: str, region: str) -> list[dict]:
    queries = [
        f"купить {material} цена поставщик {region}",
        f"{material} оптом производитель {region} цена",
        f"{material} поставщик отзывы рейтинг {region}",
    ]
    results = []
    for query in queries:
        try:
            resp = httpx.post(
                "https://google.serper.dev/search",
                headers={"X-API-KEY": SERPER_API_KEY, "Content-Type": "application/json"},
                json={"q": query, "gl": "ru", "hl": "ru", "num": 5},
                timeout=15,
            )
            data = resp.json()
            for item in data.get("organic", []):
                results.append({
                    "title": item.get("title", ""),
                    "link": item.get("link", ""),
                    "snippet": item.get("snippet", ""),
                })
        except Exception as e:
            logger.error(f"Serper error for '{material}': {e}")
    return results[:10]


def extract_materials_from_pdf(pdf_bytes: bytes) -> list[str]:
    logger.info(f"PDF size: {len(pdf_bytes)} bytes")
    pages_text = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        total = len(pdf.pages)
        max_pages = min(total, 30)
        logger.info(f"PDF opened: {total} pages, processing first {max_pages}")
        for i in range(max_pages):
            page = pdf.pages[i]
            page_text = page.extract_text() or ""
            pages_text.append(page_text)
            page.flush_cache()
            logger.info(f"Page {i + 1}/{max_pages} extracted ({len(page_text)} chars)")

    text = "\n".join(pages_text)
    logger.info(f"Total text: {len(text)} chars")
    if not text.strip():
        return []

    logger.info("Calling OpenRouter API for material extraction...")
    try:
        resp = httpx.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers=OPENROUTER_HEADERS,
            json={
                "model": "anthropic/claude-opus-4.8",
                "max_tokens": 2000,
                "messages": [{
                    "role": "user",
                    "content": f"""Из текста строительного проекта извлеки все наименования строительных материалов.
Верни ТОЛЬКО JSON-массив строк, без пояснений. Пример: ["Кирпич М150", "Цемент ПЦ 400", "Арматура А500С"]

Текст проекта:
{text[:8000]}"""
                }]
            },
            timeout=30
        )
        logger.info(f"OpenRouter response: {resp.status_code} - {resp.text[:300]}")
        resp.raise_for_status()
        data = resp.json()
        logger.info("OpenRouter API responded successfully")
    except Exception as e:
        logger.error(f"OpenRouter API error: {type(e).__name__}: {str(e)}", exc_info=True)
        raise

    raw = data["choices"][0]["message"]["content"].strip()
    try:
        if "```" in raw:
            raw = raw.split("```")[1].replace("json", "").strip()
        return json.loads(raw)
    except Exception:
        return []


def analyze_suppliers(material: str, search_results: list[dict], region: str) -> dict:
    if not search_results:
        return {
            "suppliers": [],
            "recommendation": "Данные не найдены",
            "best_supplier": "—",
        }

    snippets = "\n".join([
        f"- {r['title']}: {r['snippet']} ({r['link']})"
        for r in search_results
    ])

    resp = httpx.post(
        "https://openrouter.ai/api/v1/chat/completions",
        headers=OPENROUTER_HEADERS,
        json={
            "model": "anthropic/claude-opus-4.8",
            "max_tokens": 1500,
            "messages": [{
                "role": "user",
                "content": f"""Проанализируй результаты поиска поставщиков материала "{material}" с приоритетом региона {region}.

Результаты поиска (включая отзывы и рейтинги):
{snippets}

Верни JSON в таком формате (без пояснений):
{{
  "suppliers": [
    {{
      "name": "Название компании",
      "city": "Город/регион",
      "price": "Цена если указана или '-'",
      "contact": "Сайт или контакт",
      "reliability": "высокая/средняя/низкая",
      "review_summary": "Краткое резюме по отзывам (1 предложение) или '-' если нет данных"
    }}
  ],
  "best_supplier": "Название лучшего поставщика",
  "recommendation": "Краткая рекомендация 1-2 предложения"
}}

{region} и область ставь первыми. Максимум 5 поставщиков. Учитывай отзывы при оценке надёжности."""
            }]
        },
        timeout=30
    )
    resp.raise_for_status()
    data = resp.json()

    raw = data["choices"][0]["message"]["content"].strip()
    try:
        if "```" in raw:
            raw = raw.split("```")[1].replace("json", "").strip()
        return json.loads(raw)
    except Exception:
        return {
            "suppliers": [],
            "recommendation": raw[:200],
            "best_supplier": "—",
        }


def create_excel(materials_data: list[dict], region: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Анализ поставщиков"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    material_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    best_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    material_font = Font(bold=True, size=11)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Материал", "Поставщик", "Город/Регион", "Цена", "Контакт", "Надёжность", "Отзывы", "Рекомендация"]
    col_widths = [35, 30, 20, 15, 35, 12, 35, 40]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30

    # Регион в заголовке
    ws.cell(row=1, column=1).value = f"Материал (регион: {region})"

    row = 2
    for item in materials_data:
        material = item["material"]
        analysis = item["analysis"]
        suppliers = analysis.get("suppliers", [])
        recommendation = analysis.get("recommendation", "—")
        best = analysis.get("best_supplier", "—")

        if not suppliers:
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row=row, column=1, value=material).font = material_font
            ws.cell(row=row, column=1).fill = material_fill
            ws.cell(row=row, column=2, value="Не найдено")
            ws.cell(row=row, column=8, value=recommendation)
            row += 1
            continue

        start_row = row
        for i, s in enumerate(suppliers):
            is_best = s.get("name", "") == best
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if is_best:
                    cell.fill = best_fill

            if i == 0:
                ws.cell(row=row, column=1, value=material).font = material_font
                ws.cell(row=row, column=8, value=recommendation)

            reliability = s.get("reliability", "—")
            stars = RELIABILITY_STARS.get(reliability, reliability)

            ws.cell(row=row, column=2, value=s.get("name", "—"))
            ws.cell(row=row, column=3, value=s.get("city", "—"))
            ws.cell(row=row, column=4, value=s.get("price", "—"))
            ws.cell(row=row, column=5, value=s.get("contact", "—"))
            ws.cell(row=row, column=6, value=stars)
            ws.cell(row=row, column=7, value=s.get("review_summary", "—"))
            row += 1

        if len(suppliers) > 1:
            ws.merge_cells(f"A{start_row}:A{row - 1}")
            ws.merge_cells(f"H{start_row}:H{row - 1}")
            ws.cell(row=start_row, column=1).fill = material_fill
            ws.cell(row=start_row, column=1).alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="left"
            )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.error("Exception while handling update:", exc_info=context.error)


async def region_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.args:
        region = " ".join(context.args)
        context.user_data["region"] = region
        await update.message.reply_text(
            f"📍 Регион установлен: *{region}*\n\nТеперь поиск поставщиков будет вестись с приоритетом этого региона.",
            parse_mode="Markdown"
        )
    else:
        current = context.user_data.get("region", DEFAULT_REGION)
        await update.message.reply_text(
            f"📍 Текущий регион: *{current}*\n\n"
            f"Чтобы изменить:\n`/region Санкт-Петербург`\n`/region Краснодар`\n`/region Екатеринбург`",
            parse_mode="Markdown"
        )


async def cancel_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    cancelled_chats.add(chat_id)
    await update.message.reply_text("🛑 Отмена запрошена. Обработка остановится после текущего шага.")


async def test_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("🔄 Проверяю соединение с OpenRouter API...")
    try:
        url = "https://openrouter.ai/api/v1/chat/completions"
        payload = {
            "model": "anthropic/claude-opus-4.8",
            "max_tokens": 10,
            "messages": [{"role": "user", "content": "ping"}]
        }
        logger.info(f"Sending request to: {url}")
        logger.info(f"Headers: {list(OPENROUTER_HEADERS.keys())}")
        logger.info(f"Payload: {payload}")

        resp = await asyncio.to_thread(
            lambda: httpx.post(url, headers=OPENROUTER_HEADERS, json=payload, timeout=10)
        )
        logger.info(f"OpenRouter response status: {resp.status_code}")
        logger.info(f"OpenRouter response headers: {dict(resp.headers)}")
        logger.info(f"OpenRouter response body: {resp.text[:500]}")
        resp.raise_for_status()
        await update.message.reply_text("✅ OpenRouter API работает!")
    except Exception as e:
        logger.error(f"OpenRouter test error: {type(e).__name__}: {str(e)}", exc_info=True)
        await update.message.reply_text(f"❌ Ошибка OpenRouter API: {str(e)[:300]}")


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    region = context.user_data.get("region", DEFAULT_REGION)
    await update.message.reply_text(
        "Привет! Я помогу найти поставщиков строительных материалов.\n\n"
        "Отправь мне:\n"
        "📄 *PDF проекта* — извлеку материалы автоматически\n"
        "📝 *Список материалов текстом* — по одному на строку\n\n"
        f"📍 Текущий регион поиска: *{region}*\n"
        "Изменить: `/region Санкт-Петербург`\n\n"
        "Найду поставщиков, отсортирую по регионам и дам рекомендации с рейтингом надёжности.",
        parse_mode="Markdown"
    )


async def handle_pdf(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        region = context.user_data.get("region", DEFAULT_REGION)
        await update.message.reply_text("1️⃣ Читаю PDF...")

        file = await update.message.document.get_file()
        pdf_bytes = await file.download_as_bytearray()

        await update.message.reply_text("2️⃣ Извлекаю наименования материалов...")
        materials = await asyncio.to_thread(extract_materials_from_pdf, bytes(pdf_bytes))

        if not materials:
            await update.message.reply_text("❌ Не удалось извлечь материалы из PDF. Попробуй отправить список текстом.")
            return

        await update.message.reply_text(
            f"✅ Найдено {len(materials)} материалов:\n" +
            "\n".join(f"• {m}" for m in materials[:20]) +
            ("\n..." if len(materials) > 20 else "")
        )

        await process_materials(update, context, materials)

    except Exception as e:
        logger.error(f"handle_pdf error: {e}", exc_info=True)
        try:
            await update.message.reply_text(f"❌ Ошибка при обработке PDF: {str(e)[:300]}")
        except Exception:
            pass


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        text = update.message.text.strip()
        materials = [line.strip() for line in text.split("\n") if line.strip()]

        if not materials:
            await update.message.reply_text("Пришли список материалов — по одному на строку.")
            return

        await process_materials(update, context, materials)

    except Exception as e:
        logger.error(f"handle_text error: {e}", exc_info=True)
        try:
            await update.message.reply_text(f"❌ Ошибка: {str(e)[:300]}")
        except Exception:
            pass


async def process_materials(update: Update, context: ContextTypes.DEFAULT_TYPE, materials: list[str]) -> None:
    chat_id = update.effective_chat.id
    region = context.user_data.get("region", DEFAULT_REGION)
    cancelled_chats.discard(chat_id)
    materials_data = []

    await update.message.reply_text(
        f"4️⃣ Отправляю агентов в интернет на поиск поставщиков...\n"
        f"({len(materials)} материалов · 📍 {region})\n\n"
        f"⛔ Для отмены напиши /cancel"
    )

    for i, material in enumerate(materials, 1):
        if chat_id in cancelled_chats:
            cancelled_chats.discard(chat_id)
            await update.message.reply_text(f"🛑 Обработка отменена на материале {i}/{len(materials)}.")
            return
        await update.message.reply_text(f"🔍 [{i}/{len(materials)}] {material}")
        search_results = await asyncio.to_thread(search_suppliers, material, region)
        analysis = await asyncio.to_thread(analyze_suppliers, material, search_results, region)
        materials_data.append({"material": material, "analysis": analysis})

    await update.message.reply_text("5️⃣ Сравниваю предложения по цене, качеству и надёжности...")
    await update.message.reply_text("6️⃣ Пишу рекомендации и анализирую отзывы...")
    await update.message.reply_text("7️⃣ Сортирую поставщиков по регионам...")
    await update.message.reply_text("3️⃣ Формирую таблицу Excel...")

    excel_bytes = await asyncio.to_thread(create_excel, materials_data, region)
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"Поставщики_{region}_{date_str}.xlsx"

    await update.message.reply_document(
        document=io.BytesIO(excel_bytes),
        filename=filename,
        caption=f"8️⃣ Excel-таблица готова! 📍 Регион: {region}\n🟢 Зелёным — рекомендуемые поставщики\n⭐⭐⭐⭐⭐ — рейтинг надёжности по отзывам"
    )


async def post_init(app: Application) -> None:
    await app.bot.set_my_commands([
        BotCommand("start", "Начать работу с ботом"),
        BotCommand("region", "Установить регион поиска"),
        BotCommand("cancel", "Отменить текущую обработку"),
        BotCommand("test", "Проверить соединение с Claude API"),
    ])


def main() -> None:
    app = Application.builder().token(TELEGRAM_TOKEN).post_init(post_init).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("region", region_cmd))
    app.add_handler(CommandHandler("cancel", cancel_cmd))
    app.add_handler(CommandHandler("test", test_cmd))
    app.add_handler(MessageHandler(filters.Document.PDF, handle_pdf))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_error_handler(error_handler)
    app.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)


if __name__ == "__main__":
    main()
