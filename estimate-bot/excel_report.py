"""Export EstimateResult to Excel (.xlsx)."""

import logging
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from calculator import EstimateResult

logger = logging.getLogger(__name__)


def export_to_excel(estimate: EstimateResult, region: str) -> str:
    """
    Export verified estimate to .xlsx file.
    Returns file path.
    Raises RuntimeError if not verified.
    """
    if not estimate.verified:
        raise RuntimeError("Refuse to export unverified estimate")

    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    # Header: region, date
    ws["A1"] = f"Смета объекта ({region})"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

    # Column headers
    headers = ["#", "Тип", "Наименование", "Ед.", "Кол-во", "Цена ед., ₽", "Сумма, ₽", "Диапазон*", "В итоге?", "Уверенность", "Источник"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for idx, line in enumerate(estimate.lines, start=1):
        row = 4 + idx

        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = "Материал" if line.type == "material" else "Работа"
        ws.cell(row=row, column=3).value = line.name
        ws.cell(row=row, column=4).value = line.unit
        ws.cell(row=row, column=5).value = float(line.qty) if line.qty else None
        ws.cell(row=row, column=6).value = float(line.unit_price) if line.unit_price else None
        ws.cell(row=row, column=7).value = float(line.line_sum) if line.line_sum else None

        # Range column
        if line.sum_min and line.sum_max:
            ws.cell(row=row, column=8).value = f"{float(line.sum_min):.2f} - {float(line.sum_max):.2f}"
        else:
            ws.cell(row=row, column=8).value = None

        ws.cell(row=row, column=9).value = "Да" if line.included_in_total else "Нет"
        ws.cell(row=row, column=10).value = line.confidence
        ws.cell(row=row, column=11).value = line.source

    # Summary section
    summary_row = 5 + len(estimate.lines) + 2
    ws.cell(row=summary_row, column=1).value = "Материалы:"
    ws.cell(row=summary_row, column=2).value = float(estimate.total_materials)
    ws.cell(row=summary_row, column=2).font = Font(bold=True)

    ws.cell(row=summary_row + 1, column=1).value = "Работы:"
    ws.cell(row=summary_row + 1, column=2).value = float(estimate.total_works)
    ws.cell(row=summary_row + 1, column=2).font = Font(bold=True)

    ws.cell(row=summary_row + 2, column=1).value = "ИТОГО:"
    ws.cell(row=summary_row + 2, column=2).value = float(estimate.grand_total)
    ws.cell(row=summary_row + 2, column=2).font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 2, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    if estimate.verified:
        ws.cell(row=summary_row + 3, column=1).value = "✓ Расчёт перепроверен калькулятором"
        ws.cell(row=summary_row + 3, column=1).font = Font(italic=True, color="008000")

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 8
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 20

    # Save
    filename = f"smeta_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    logger.info(f"Excel report saved: {filename}")
    return filename
